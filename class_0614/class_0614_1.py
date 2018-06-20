

#1：利用6-12号写的excel数据读取类以及6月9号写的HTTP单元测试类，完成老黄历的单元测试，且输出测试报告。
# （注意是要从Excel里面获取测试 数据哦~并且把测试结果写回到Excel中去。

from openpyxl import load_workbook


class ExcelCase:
    def __init__(self, work_book, sheet):
        self.work_book = load_workbook(work_book)
        self.sheet = self.work_book[sheet]
        self.list_2 = []

    #把单元测试报告里面的值写入到Excel中
    def write_testcase(self):
        self.sheet.cell(4, 7).value = '错误的请求KEY!!'
        self.sheet.cell(4, 8).value = 'fail'
        self.sheet.cell(4, 9).value = 'ft1_1: 开始执行测试用例结束测试Traceback (most recent call last):File "/Users/mantis/Documents/Python/Test_demo/class_0614/class_0614_2.py", line 22, in test_get_post_ruquestsself.assertEqual(错误的请求KEY!!,response[reason])AssertionError: 错误的请求KEY!!!= successed- 错误的请求KEY!!+ successed'
        self.work_book.save('test_case_2.xlsx')

    #把Excel的值读取出来
    def read_testcase(self):
        for i in range(2, self.sheet.max_row + 1):
            list_1 = []
            for j in range(1, self.sheet.max_column + 1):
                list_1.append(self.sheet.cell(i, j).value)
            self.list_2.append(list_1)
        return self.list_2


if __name__ == '__main__':
    t = ExcelCase('test_case_2.xlsx', 'test')
    print(t.read_testcase())
    t.write_testcase()



#2：结合mysqlconnector 编写一个数据库操作类，完成指定数据库的操作类。


