

import pymysql.connections

#1：利用6-12号写的excel数据读取类以及6月9号写的HTTP单元测试类，完成老黄历的单元测试，且输出测试报告。
# （注意是要从Excel里面获取测试 数据哦~并且把测试结果写回到Excel中去。

#测试数据/结果--Excel里面
#excel读取类--专门负责数据的读取/存储
#HTTP请求类--专门完成HTTP请求
#HTTP请求测试类--单元测试
#最终输出测试报告HTML

from openpyxl import load_workbook

class DoExcel:
    def __init__(self, work_excel, sheet):
        self.work_excel = work_excel
        self.sheet = sheet

    def write_data(self,row,actaul,result):      #写入测试结果
        wb=load_workbook(self.work_excel)
        sheet=wb[self.sheet]
        sheet.cell(row,7).value=actaul
        sheet.cell(row,8).value=result
        wb.save(self.work_excel)


    def read_data(self):
        wd = load_workbook(self.work_excel)
        sheet = wd[self.sheet]
        test_data = []
        for row in range(2, sheet.max_row + 1):

            list_1 = []
            for i in range(1, sheet.max_column + 1):
                list_1.append(sheet.cell(row, i).value)
            test_data.append(list_1)
        return test_data


if __name__ == '__main__':
    t = DoExcel('test_case_0614.xlsx', 'test_data')
    print(t.read_data())



#2：结合mysqlconnector 编写一个数据库操作类，完成指定数据库的操作类。

class DoMysql:
    def __init__(self,config,sql,data):
        self.config=config
        self.sql=sql
        self.data=data

    def get_data(self):
        cnn=pymysql.connect(**self.config)
        cursor=cnn.cursor()
        cursor.execute(self.sql,self.data)
        result=cursor.fetchall()
        cursor.close()
        cnn.close()
        return result
if __name__ == '__main__':
    config = {'host': '118.126.108.173',  # 主机
              'user': 'python',  # 用户名
              'password': 'python5666',  # 密码
              'port': 3306,  # 端口
              'database': 'test_summer',  # 数据库名
              }

