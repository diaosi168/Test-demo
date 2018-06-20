
#Excel workbook 工作簿-->sheet 表单--》cell单元格
from openpyxl import load_workbook #openpyxl  xlsx后缀的Excel

#1.打开工作簿
work_book=load_workbook('test001.xlsx')

#2.确定表单名 sheet
sheet=work_book['test']

#3.确定单元格的位置 cell
print(sheet.cell(2,6).value)
print(sheet.cell(3,5).value)

print(sheet.max_row)
print(sheet.max_column)
#4.写入值
sheet.cell(2,7).value='duolala'
work_book.save('test_case_2.xlsx')


