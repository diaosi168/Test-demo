

#1：url地址：http://v.juhe.cn/laohuangli/d
# 请求参数：{'date':'2018-09-11','key':'a8f2732319cf0ad3cce8ec6ef7aa4f33'}

#2：请根据上课内容对老黄历编写至少5条用例，上述给出的数据都是正确的请求数据

#3：把用例的数据写到Excel里面，然后请把每一行的数据存到一个子列表里面，所有的行数据都放到一个大的列表里面。

#4：读取数据的的功能，请写成一个类
#from unittest.test import test_case

from openpyxl import load_workbook

class ExcelCase:
    def __init__(self,work_book,sheet):
        self.work_book = load_workbook(work_book)
        self.sheet = self.work_book[sheet]
        self.list_2 = []

    def read_testcase(self):
        for i in range(2,self.sheet.max_row+1):
            list_1=[]
            for j in range(1,self.sheet.max_column+1):
                list_1.append(self.sheet.cell(i,j).value)
            self.list_2.append(list_1)
        return self.list_2

if __name__ == '__main__':
    #work_book=load_workbook('test_case_2.xlsx')
    #sheet=work_book['test']
    t=ExcelCase('test_case_2.xlsx','test')
    print(t.read_testcase())


'''class ExcelCase:
    def __init__(self, work_excel, sheet):
        self.work_excel = work_excel
        self.sheet = sheet

    def test_date(self):
        wd = load_workbook(self.work_excel)
        sheet = wd[self.sheet]
        lsit_data = []
        for row in range(2, sheet.max_row + 1):

            list_1 = []
            for i in range(1, sheet.max_column + 1):
                list_1.append(sheet.cell(row, i).value)
                lsit_data.append(list_1)
        return lsit_data


if __name__ == '__main__':
    t = ExcelCase('test_case_2.xlsx', 'test')
    print(t.test_date())'''