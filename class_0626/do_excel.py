
#把测试数据从excel读取出来

from openpyxl import load_workbook

class DoExcel:
    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name
    def read_data(self):
        wb=load_workbook(self.file_path)
        sheet=wb[self.sheet_name]
        # 每一行数据存在一个列表里面，然后所有行的数据存在大列表里面
        test_data = []  # 存储所有行的数据
        for i in range(2, sheet.max_row + 1):
            sub_data = []
            for j in range(1, 8):
                sub_data.append(sheet.cell(i, j).value)
            test_data.append(sub_data)
        return (test_data)

    def write_data(self,row,actual,result):
        wb=load_workbook(self.file_path)
        sheet=wb[self.sheet_name]
        sheet.cell(row,8).value=actual
        sheet.cell(row,9).value=result

        wb.save(self.file_path)  #保存数据
if __name__ == '__main__':
    test_data=DoExcel('tase_case.xlsx','test_data').read_data()
    print(test_data)

